from datetime import datetime

from django.http import HttpResponse
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl import Workbook


def get_mileage_headers():
    headers = [
        'date',
        'postcodes',
        'distance',
    ]
    return headers


def get_mileage_row(data):

    row = [
        data.date,
        data.postcodes,
        data.distance,
    ]
    return row


class ExcellExporter():
    def __init__(self, queryset, headers, export_object):
        self.queryset = queryset
        self.workbook = Workbook()
        self.headers = headers
        self.export_object = export_object

    def get_virtual_workbook(self):
        return save_virtual_workbook(self.workbook)

    def excel_response(self):
        response = HttpResponse(
            content=self.get_virtual_workbook(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheet.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-{object}-tank.xlsx'.format(
            date=datetime.now().strftime('%Y-%m-%d'),
            object=self.export_object
        )
        return response

    def export_worksheet(self):
        worksheet = self.workbook.active

        row_num = 1
        for col_num, column_title in enumerate(self.headers, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        if self.export_object == 'mileage':
            worksheet.title = 'Mileage'
            for data in self.queryset:
                row_num += 1
                row = get_mileage_row(data)
                for col_num, cell_value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = cell_value
            # self.set_column_widths(worksheet)
            return self.excel_response()
